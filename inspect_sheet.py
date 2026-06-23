import openpyxl

wb = openpyxl.load_workbook("Tabela Revisão Programada FIAT Junho 2026.xlsx", data_only=True)

print("Preços de Mão de Obra e óleos em Junho 2026:")
for sheetname in wb.sheetnames[:10]: # primeiras 10 abas
    sheet = wb[sheetname]
    for r in range(1, sheet.max_row + 1):
        desc = sheet.cell(row=r, column=2).value
        pn = sheet.cell(row=r, column=3).value
        preco = sheet.cell(row=r, column=4).value
        
        if desc is not None:
            desc_lower = str(desc).lower()
            if "mão-de-obra" in desc_lower or "mão de obra" in desc_lower or (pn and "mo fiat" in str(pn).lower()):
                print(f"[{sheetname}] Mão de Obra: PN={pn}, Preço={preco}")
            if "óleo" in desc_lower or "oleo" in desc_lower:
                if preco is not None:
                    print(f"[{sheetname}] Óleo: Desc='{desc}', PN={pn}, Preço={preco}")
