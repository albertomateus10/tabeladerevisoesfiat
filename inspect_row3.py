import openpyxl

wb = openpyxl.load_workbook("Tabela Revisão Programada FIAT Março 2026.xlsx", data_only=True)
sheet = wb['ARGO 1.3']

print("--- LINHA 3 COMPLETA EM MARÇO ---")
row_vals = [sheet.cell(row=3, column=c).value for c in range(1, sheet.max_column + 1)]
print(row_vals)
